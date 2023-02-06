import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
import customtkinter
from utils import create_workbook
from correct_data import *
from pathlib import Path
from variables import CASE_VARS, CLASS_VARS, MAHA_VARS, ABCDC_VARS


customtkinter.set_appearance_mode("System")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("green")  # Themes: "blue" (standard), "green", "dark-blue"


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()

        # configure window
        self.title("Data Migration Assistant")
        self.geometry(f"{980}x{580}")

        # configure grid layout (4x4)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure((2), weight=0)
        self.grid_rowconfigure((0, 1, 2), weight=1)

        # Create Initial workbook to save all data to
        self.workbook = create_workbook()
        self.template_client_sheet = self.workbook['Client Case Session Upload']
        self.template_class_sheet = self.workbook['Client Class Upload']

        # Application Variables
        self.AGENCY = tk.StringVar(value='TEMPLATE')
        self.HAS_CLIENTS = tk.StringVar(value=str(False))
        self.HAS_CLASSES = tk.StringVar(value=str(False))
        self.FIX_ADDRESS = tk.StringVar(value=str(False))
        self.FIX_ZIPCODE = tk.StringVar(value=str(False))
        self.SAVE_AS = tk.StringVar(value="")
        self.CLASS_VARS = CASE_VARS
        self.agency_filename = ""
        self.agency_data = []
        self.clients = {}
        self.classes = {}
        self.clients_idx = 1
        self.classes_idx = 1

        # create sidebar frame with widgets
        self.sidebar_frame = customtkinter.CTkFrame(self, width=140, corner_radius=0)
        self.sidebar_frame.grid(row=0, column=0, rowspan=4, sticky="nsew")
        self.sidebar_frame.grid_rowconfigure(4, weight=1)
        self.logo_label = customtkinter.CTkLabel(self.sidebar_frame, text="Migration Assistant", font=customtkinter.CTkFont(size=20, weight="bold"))
        self.logo_label.grid(row=0, column=0, padx=20, pady=(20, 10))
        self.open_button = customtkinter.CTkButton(self.sidebar_frame, command=self.browse_file, text="Open Spreadsheet")
        self.open_button.grid(row=1, column=0, padx=20, pady=10)
        self.process_clients_button = customtkinter.CTkButton(self.sidebar_frame, command=self.process_clients, text="Process Clients")
        self.process_clients_button.grid(row=2, column=0, padx=20, pady=10)
        self.process_classes_button = customtkinter.CTkButton(self.sidebar_frame, command=self.process_classes, text="Process Classes")
        self.process_classes_button.grid(row=3, column=0, padx=20, pady=10)

        # Appearance Mode Switcher
        self.appearance_mode_label = customtkinter.CTkLabel(self.sidebar_frame, text="Appearance Mode:", anchor="w")
        self.appearance_mode_label.grid(row=5, column=0, padx=20, pady=(10, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode_event)
        self.appearance_mode_optionemenu.grid(row=6, column=0, padx=20, pady=(10, 10))

        # UI Scaling Switcher
        self.scaling_label = customtkinter.CTkLabel(self.sidebar_frame, text="UI Scaling:", anchor="w")
        self.scaling_label.grid(row=7, column=0, padx=20, pady=(10, 0))
        self.scaling_optionemenu = customtkinter.CTkOptionMenu(self.sidebar_frame, values=["80%", "90%", "100%", "110%", "120%"],
                                                               command=self.change_scaling_event)
        self.scaling_optionemenu.grid(row=8, column=0, padx=20, pady=(10, 20))

        # create main entry and button
        self.entry = customtkinter.CTkEntry(self, placeholder_text="Save File As...", textvariable=self.SAVE_AS)
        self.entry.grid(row=3, column=1, columnspan=2, padx=(20, 0), pady=(20, 20), sticky="nsew")

        self.download_button = customtkinter.CTkButton(master=self, fg_color="transparent", border_width=2, text_color=("gray10", "#DCE4EE"), text="Download", command=self.save_workbook)
        self.download_button.grid(row=3, column=3, padx=(20, 20), pady=(20, 20), sticky="nsew")

        # create textbox
        self.textbox = customtkinter.CTkTextbox(self, width=250)
        self.textbox.grid(row=0, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")

        # create tabview
        self.tabview = customtkinter.CTkTabview(self, width=250)
        self.tabview.grid(row=2, column=1, padx=(20, 0), pady=(20, 0), sticky="nsew")
        self.tabview.add("Select Agency")

        self.tabview.tab("Select Agency").grid_columnconfigure(0, weight=1)

        self.agency_menu = customtkinter.CTkOptionMenu(self.tabview.tab("Select Agency"), dynamic_resizing=False,
                                                        values=["TEMPLATE", "MAHA", "FEA", "ABCDC"], command=self.agency_menu_callback)
        self.agency_menu.grid(row=0, column=0, padx=20, pady=(20, 10))


        # create upload details frame
        self.upload_details_frame = customtkinter.CTkFrame(self, width=250)
        # self.upload_details_frame.grid_columnconfigure(0, weight=1)
        self.upload_details_frame.grid(row=0, column=2, padx=(20, 20), pady=(20, 10), sticky="nsew")
        
        self.label_checkbox_group = customtkinter.CTkLabel(master=self.upload_details_frame, text="Upload Details:")
        self.label_checkbox_group.grid(row=0, column=2, columnspan=1, padx=10, pady=10, sticky="")
        self.check_box_1 = customtkinter.CTkCheckBox(master=self.upload_details_frame, variable=self.HAS_CLIENTS, text="Client Upload", onvalue="True", offvalue="False", command=self.set_checkbox_clients)
        self.check_box_1.grid(row=1, column=2, pady=10, padx=20, sticky="n")
        self.check_box_2 = customtkinter.CTkCheckBox(master=self.upload_details_frame, variable=self.HAS_CLASSES, text="Class Upload", onvalue="True", offvalue="False", command=self.set_checkbox_class)
        self.check_box_2.grid(row=2, column=2, pady=10, padx=20, sticky="n")
        self.check_box_3 = customtkinter.CTkCheckBox(master=self.upload_details_frame, variable=self.FIX_ADDRESS, text="Fix Address", onvalue="True", offvalue="False", command=self.set_checkbox_address)
        self.check_box_3.grid(row=3, column=2, pady=10, padx=20, sticky="n")

        # set default values
        self.appearance_mode_optionemenu.set("Dark")
        self.scaling_optionemenu.set("100%")
        self.agency_menu.set("Select Agency")
        self.download_button.configure(state="Disabled")
        self.process_clients_button.configure(state="Disabled")
        self.process_classes_button.configure(state="Disabled")

        self.textbox.insert("0.0", "How To For Dummies\n\n" + "Open the desired XLSX spreadsheet by clicking on the green Open Spreadsheet Button.\n\nThe Process Clients, and Classes Buttons are tied to the Upload details, if they are off in the details, you cannot click the button.\n\nBe sure to select an agency from the agency dropdown if you are uploading a specialty agency that pays monthly for uploads, otherwise select 'TEMPLATE' to show you are uploading from our exact template\n\nOnly Process Clients, or Classes one at a time for now, once you have clicked the process button for whichever spreadsheet you need, enter a filename to save it under and press the download button.\n\nCheck your download folder to ensure the data is correct, if something is wrong look over your original spreadsheet and ensure that it meets our actual templates standards.")


    def change_appearance_mode_event(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)


    def change_scaling_event(self, new_scaling: str):
        new_scaling_float = int(new_scaling.replace("%", "")) / 100
        customtkinter.set_widget_scaling(new_scaling_float)


    def set_checkbox_clients(self):
        self.process_clients_button.configure(state=self.get_state(self.HAS_CLIENTS.get()))


    def set_checkbox_class(self):
        self.process_classes_button.configure(state=self.get_state(self.HAS_CLASSES.get()))
        
    
    def get_state(self, bool):
        if bool:
            return "Active"

        return "Disabled"


    def set_checkbox_address(self):
        self.FIX_ADDRESS.set(not eval(self.FIX_ADDRESS.get()))


    def browse_file(self):
        self.agency_filename = filedialog.askopenfilename()
        self.agency_data = load_workbook(filename=self.agency_filename, data_only=True)


    def agency_menu_callback(self, choice):
        self.set_agency_vars()
        self.AGENCY.set(choice)

    
    def set_agency_vars(self):
        if self.AGENCY.get() == "MAHA":
            self.CLASS_VARS = MAHA_VARS

        if self.AGENCY.get() == "ABCDC":
            self.CLASS_VARS = ABCDC_VARS

        self.CLASS_VARS = CLASS_VARS


    def process_clients(self):
        if eval(self.HAS_CLIENTS.get()):
            # Process the data here
            for row in self.agency_data.active.iter_rows(min_row=CASE_VARS["STARTINGROW"], values_only=True, min_col=CASE_VARS["STARTINGCOL"]):
                client_id = row[CASE_VARS["CLIENTID"]]
                client = {
                    'ID': self.clients_idx,
                    'clientId': correct_legacy_id(client_id),
                    'FirstName': row[CASE_VARS["FIRSTNAME"]],
                    'LastName': row[CASE_VARS["LASTNAME"]],
                    'Phone': clean_phone(row[CASE_VARS["PHONE"]]),
                    'AddressNumber': row[CASE_VARS["ADDRESSNUMBER"]],
                    'StreetName': row[CASE_VARS["STREETNAME"]],
                    'ClientCity': row[CASE_VARS["CITY"]],
                    'ClientState': row[CASE_VARS["STATE"]],
                    'ClientCounty': row[CASE_VARS["COUNTY"]],
                    'Zip': row[CASE_VARS["ZIP"]],
                    'Race': correct_race(row[CASE_VARS["RACE"]]),
                    'Ethnicity': correct_ethnicity(row[CASE_VARS["ETHNICITY"]]),
                    'EnglishProficiencyLevel': correct_eng_prof(row[CASE_VARS["ENGLISHPROF"]]),
                    'HouseholdIncome': row[CASE_VARS["HOUSEHOLDINCOME"]],
                    'HouseholdSize': row[CASE_VARS["HOUSEHOLDSIZE"]],
                    'ApartmentNumber': row[CASE_VARS["APTNUMBER"]],
                    'Email': row[CASE_VARS["EMAIL"]],
                    'InitialCaseType': correct_case_type(row[CASE_VARS["INITIALCASETYPE"]]),
                    'DateOfBirth': correct_date(row[CASE_VARS["DATEOFBIRTH"]]),
                    'CounselorName': row[CASE_VARS["COUNSELOR"]],
                    'IntakeDate': correct_date(row[CASE_VARS["INTAKEDATE"]]),
                    'CaseType': correct_case_type(row[CASE_VARS["CASETYPE"]]),
                    'CaseStartDate': correct_date(row[CASE_VARS["CASESTARTDATE"]]),
                    'CaseID': correct_case_id(row[CASE_VARS["CASEID"]]),
                    'SessionID': row[CASE_VARS["SESSIONID"]],
                    'Date': row[CASE_VARS["DATE"]],
                    'NOFAGrant': correct_nofa(row[CASE_VARS["NOFAGRANT"]]),
                    '9a': None,
                    '9b': None,
                    '9c': None,
                    '9d': None,
                    '9e': None,
                    '9f': None,
                    '10a': None,
                    '10b': None,
                    '10c': None,
                    '10d': None,
                    '10e': None,
                    '10f': None,
                    '10g': None,
                    '10h': None,
                    '10i': None,
                    '10j': None,
                    '10k': None,
                    '10l': None,
                    '10m': None,
                    '10n': None,
                    'SessionNotes': row[CASE_VARS["NOTES"]],
                    'ruralareastatus': correct_rural(row[CASE_VARS["RURALSTATUS"]]),
                    'HouseholdType': correct_household(row[CASE_VARS["HOUSEHOLDTYPE"]])
                }

                # Split the address
                if eval(self.FIX_ADDRESS.get()):
                    split_address(row[CASE_VARS["STREETNAME"]], client)

                if eval(self.FIX_ZIPCODE.get()):
                    get_zipcode(row[CASE_VARS["CITY"]], row[CASE_VARS["STATE"]], client)

                # Save the client by the ID for easy Access
                self.clients[self.clients_idx] = client
                self.clients_idx += 1

            # THIS IS DONE LAST FOR THE CLIENT SHEET
            # GRAB ALL DATA BEFORE THIS UNLESS TESTING
            # Add each client to the new spreadsheet
            for client in self.clients.values():
                self.template_client_sheet.append(list(client.values()))

            self.download_button.configure(state="Active")

        return self.agency_data


    def process_classes(self):
        if eval(self.HAS_CLASSES.get()):
            # Process the data here
            # Start creating initial hash table of classes
            for row in self.agency_data.active.iter_rows(min_row=self.CLASS_VARS["STARTINGROW"], values_only=True, min_col=self.CLASS_VARS["STARTINGCOL"]):
                client_id = row[self.CLASS_VARS["CLIENTID"]]
                # classDateFix = correct_date(row[17])
                client = {
                    'ID': self.classes_idx,
                    'clientId': correct_legacy_id(client_id),
                    'FirstName': row[self.CLASS_VARS["FIRSTNAME"]],
                    'LastName': row[self.CLASS_VARS["LASTNAME"]],
                    'Phone': clean_phone(row[self.CLASS_VARS["PHONE"]]),
                    'AddressNumber': row[self.CLASS_VARS["ADDRESSNUMBER"]],
                    'StreetName': row[self.CLASS_VARS["STREETNAME"]],
                    'ClientCity': row[self.CLASS_VARS["CITY"]],
                    'ClientState': row[self.CLASS_VARS["STATE"]],
                    'ClientCounty': row[self.CLASS_VARS["COUNTY"]],
                    'Zip': row[self.CLASS_VARS["ZIP"]],
                    'Race': correct_race(row[self.CLASS_VARS["RACE"]]),
                    'Ethnicity': correct_ethnicity(row[self.CLASS_VARS["ETHNICITY"]]),
                    'EnglishProficiencyLevel': correct_eng_prof(row[self.CLASS_VARS["ENGLISHPROF"]]),
                    'HouseholdIncome': row[self.CLASS_VARS["HOUSEHOLDINCOME"]],
                    'HouseholdSize': row[self.CLASS_VARS["HOUSEHOLDSIZE"]],
                    'ApartmentNumber': row[self.CLASS_VARS["APTNUMBER"]],
                    'Email': row[self.CLASS_VARS["EMAIL"]],
                    'InitialCaseType': correct_case_type(row[self.CLASS_VARS["INITIALCASETYPE"]]),
                    'DateOfBirth': correct_date(row[self.CLASS_VARS["DATEOFBIRTH"]]),
                    'CounselorName': row[self.CLASS_VARS["COUNSELOR"]],
                    'IntakeDate': correct_date(row[self.CLASS_VARS["INTAKEDATE"]]),
                    'CourseID': row[self.CLASS_VARS["COURSEID"]],
                    'ClassDate': row[self.CLASS_VARS["CLASSDATE"]],
                    'AttendanceStatus': correct_attendance(row[self.CLASS_VARS["ATTENDANCESTATUS"]]),
                    'RuralAreaStatus': correct_rural(row[self.CLASS_VARS["RURALSTATUS"]]),
                    'HouseholdType': correct_household(row[self.CLASS_VARS["HOUSEHOLDTYPE"]])
                }

                if eval(self.FIX_ADDRESS.get()):
                    split_address(row[self.CLASS_VARS["STREETNAME"]], client)

                if eval(self.FIX_ZIPCODE.get()):
                    get_zipcode(row[self.CLASS_VARS["CITY"]], row[self.CLASS_VARS["STATE"]], client)

                # Save the client by the ID for easy Access
                self.classes[self.classes_idx] = client
                self.classes_idx += 1

            for client in self.classes.values():
                self.template_class_sheet.append(list(client.values()))

            self.download_button.configure(state="Active")

        return self.agency_data


    def save_workbook(self):
        download_path = str(Path.home() / "Downloads")
        outputFileName = f"{download_path}\\{self.SAVE_AS.get()}.xlsx"
        self.workbook.save(filename=outputFileName)
        self.reset_assistant()


    def reset_assistant(self):
        self.agency_filename = ""
        self.agency_data = ""
        self.download_button.configure(state="Disabled")
        self.process_clients_button.configure(state="Disabled")
        self.process_classes_button.configure(state="Disabled")
        self.agency_menu.set("Select Agency")
        self.AGENCY.set("TEMPLATE")
        self.HAS_CLIENTS.set("False")
        self.HAS_CLASSES.set("False")
        self.FIX_ADDRESS.set("False")
        self.SAVE_AS.set("")
        self.CLASS_VARS = CLASS_VARS
        self.clients = {}
        self.classes = {}
        self.clients_idx = 1
        self.classes_idx = 1


if __name__ == "__main__":
    app = App()
    app.mainloop()