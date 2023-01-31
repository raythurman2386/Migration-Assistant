from openpyxl import Workbook


def create_workbook():
    # *****************************************************************************************
    # Open new workbook, create tabs, and write headers
    # Full template workbook setup and initialization
    workbook = Workbook()
    # workbook.iso_dates = True
    template_client_sheet = workbook.create_sheet('Client Case Session Upload')
    template_case_sheet = workbook.create_sheet('Client Class Upload')

    # Headers for all tabs of the workbook
    client_case_session_upload_header = ['ID', 'ClientID', 'FirstName', 'LastName', 'Phone', 'AddressNumber', 'StreetName', 'City', 'State', 'County',
                                         'Zip', 'Race', 'Ethnicity', 'EnglishProficiencyLevel', 'HouseholdIncome',
                                         'HouseholdSize', 'ApartmentNumber', 'Email', 'InitialCaseType', 'DateofBirth', 'CounselorName', 'IntakeDate', 'CaseType', 'CaseStartDate', 'CaseID',
                                         'SessionID', 'Date', 'NOFAGrant', '9a', '9b', '9c', '9d', '9e', '9f', '10a', '10b', '10c', '10d',
                                         '10e', '10f', '10g', '10h', '10i', '10j', '10k', '10l', '10m', '10n', 'SessionNotes', 'RuralAreaStatus', 'HouseholdType']
    client_class_upload_header = ['ID', 'ClientID', 'FirstName', 'LastName', 'Phone', 'AddressNumber', 'StreetName', 'City', 'State', 'County',
                                  'Zip', 'Race', 'Ethnicity', 'EnglishProficiencyLevel', 'HouseholdIncome',
                                  'HouseholdSize', 'ApartmentNumber', 'Email', 'InitialCaseType', 'DateofBirth', 'CounselorName', 'IntakeDate',
                                  'ClassNO', 'ClassDate', 'AttendanceStatus', 'RuralAreaStatus', 'HouseholdType']

    # Append our proper header to the new worksheet
    template_client_sheet.append(client_case_session_upload_header)
    template_case_sheet.append(client_class_upload_header)
    # *****************************************************************************************
    workbook.remove(workbook.active)
    return workbook


options = {
    'race': {
        'a': 'American Indian or Alaskan Native',
        'b': 'Asian',
        'c': 'Black or African American',
        'd': 'Chose not to respond',
        'e': 'Native Hawaiian or other Pacific Islander',
        'f': 'More than one race',
        'g': 'White'
    },
    'ethnicity': {
        'a': 'Hispanic',
        'b': 'Not Hispanic',
        'c': 'Chose not to respond'
    },
    'englishProficiency': {
        'a': 'Limited English Proficient',
        'b': 'Not Limited English Proficient',
        'c': 'Chose not to respond'
    },
    'caseType': {
        'a': 'Home Purchase',
        'b': 'Seeking Shelter',
        'c': 'Home Owner Services',
        'd': 'Mortgage Modification',
        'e': 'Rental topics',
        'f': 'Education',
        'g': 'Other - Non HUD Reportable'
    },
    'NOFA': {
        'a': 'NOFA 2017-1 COMP',
        'b': 'NOFA 2019-1 COMP',
        'c': 'NOFA 2020-1 COMP',
        'd': 'NOFA 2021-1 COMP'
    },
    'Attended': {
        'a': 'Attended',
        'b': 'No Show',
        'c': 'Enrolled'
    },
    'RuralAreaStatus': {
        'a': 'Lives in a rural area',
        'b': 'Does not live in a rural area',
        'c': 'Chose not to respond'
    },
    'HouseholdType': {
        'a': 'Single adult',
        'b': 'Female-headed single parent household',
        'c': 'Male-headed single parent household',
        'd': 'Married without dependents',
        'e': 'Married with dependents',
        'f': 'Two or more unrelated adults',
        'g': 'Other'
    },
    'race5DEPRECATED': {
        'a': 'American Indian/Alaskan Native',
        'b': 'Asian',
        'c': 'Black or African American',
        'd': 'Native Hawaiian or Other Pacific Islander',
        'e': 'White',
        'f': 'American Indian or Alaska Native and White',
        'g': 'Asian and White',
        'h': 'Black or African American and White',
        'i': 'American Indian or Alaska Native and Black or African American',
        'j': 'Other multiple race',
        'k': 'Chose not to respond'
    },
}
