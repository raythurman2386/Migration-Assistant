from openpyxl import load_workbook
from utils import create_workbook
from tkinter import filedialog
from correct_data import *
from pathlib import Path
from variables import CASE_VARS, CLASS_VARS, MAHA_VARS, ABCDC_VARS
import tkinter as tk
import customtkinter

# Tkinter setup
# System Settings
customtkinter.set_appearance_mode("System")
customtkinter.set_default_color_theme("green")

# Define App Frame
root = customtkinter.CTk()
root.geometry("720x580")
root.title("Data Migration Assistant")

# Create Initial workbook to save all data to
workbook = create_workbook()
template_client_sheet = workbook['Client Case Session Upload']
template_class_sheet = workbook['Client Class Upload']
AGENCY = tk.StringVar(value='TEMPLATE')
HAS_CLIENTS = tk.StringVar(value=str(False))
HAS_CLASSES = tk.StringVar(value=str(False))
FIX_ADDRESS = tk.StringVar(value=str(False))
FIX_ZIPCODE = tk.StringVar(value=str(False))
SAVE_FILE_AS = tk.StringVar(value="Save As . . .")
agency_filename = ""
agency_data = ""

# UI Elements
title = customtkinter.CTkLabel(root, text="Migration Assitant", font=('Arial', 16))
title.pack(padx=10, pady=10)

# Agency Selector Combo Box
agency_frame = customtkinter.CTkFrame(master=root, width=680, height=200, corner_radius=10)
agency_frame.pack(padx=20, pady=20)

def agency_menu_callback(choice):
    AGENCY.set(choice)


agency_menu_label = customtkinter.CTkLabel(agency_frame, text="Please select if uploading a special agency.", font=('Arial', 14))
agency_menu_label.pack(padx=10, pady=10, side=tk.LEFT)

agency_menu = customtkinter.CTkOptionMenu(master=agency_frame,values=['TEMPLATE', 'MAHA', 'FEA', 'ABCDC'], command=agency_menu_callback)
agency_menu.pack(padx=10, pady=10, side=tk.RIGHT)
agency_menu.set(AGENCY.get())


# checkbox frame to control app variables
def set_checkbox():
    # print(f"System Variables: Clients: {HAS_CLIENTS.get()}, Classes: {HAS_CLASSES.get()}, Address: {FIX_ADDRESS.get()}, Zipcode: {FIX_ZIPCODE.get()}")
    pass

checkbox_frame = customtkinter.CTkFrame(master=root, width=680, height=200, corner_radius=10)
checkbox_frame.pack(padx=20, pady=20)

checkbox_title = customtkinter.CTkLabel(master=checkbox_frame, text="Select migration details.", font=('Arial', 14))
checkbox_title.pack(padx=10, pady=10, side=tk.TOP)

client_checkbox = customtkinter.CTkCheckBox(master=checkbox_frame, text="Has Clients", command=set_checkbox, variable=HAS_CLIENTS, onvalue="True", offvalue="False")
client_checkbox.pack(padx=10, pady=10, side=tk.LEFT)

class_checkbox = customtkinter.CTkCheckBox(master=checkbox_frame, text="Has Classes", command=set_checkbox, variable=HAS_CLASSES, onvalue="True", offvalue="False")
class_checkbox.pack(padx=10, pady=10, side=tk.LEFT)

address_checkbox = customtkinter.CTkCheckBox(master=checkbox_frame, text="Fix Address", command=set_checkbox, variable=FIX_ADDRESS, onvalue="True", offvalue="False")
address_checkbox.pack(padx=10, pady=10, side=tk.RIGHT)

zip_checkbox = customtkinter.CTkCheckBox(master=checkbox_frame, text="Fix Zipcode", command=set_checkbox, variable=FIX_ZIPCODE, onvalue="True", offvalue="False")
zip_checkbox.pack(padx=10, pady=10, side=tk.RIGHT)


def reset_assistant():
    # Create Initial workbook to save all data to
    workbook = create_workbook()
    template_client_sheet = workbook['Client Case Session Upload']
    template_class_sheet = workbook['Client Class Upload']
    AGENCY.set('TEMPLATE')
    HAS_CLIENTS.set(str(False))
    HAS_CLASSES.set(str(False))
    FIX_ADDRESS.set(str(False))
    FIX_ZIPCODE.set(str(False))
    SAVE_FILE_AS.set("")
    # instantiate variables
    agency_filename = ""
    agency_data = ""


# Save final outcome of data manipulation in Downloads Folder
def save_workbook():
    # Download to the downloads folder
    download_path = str(Path.home() / "Downloads")
    # Save the worksheet when all is complete
    outputFileName = f'{download_path}\\final_{SAVE_FILE_AS.get()}{AGENCY.get()}.xlsx'
    workbook.save(filename=outputFileName)
    reset_assistant()    


def process_clients(client_data):
    clients = {}
    clients_idx = 1
    for row in client_data.active.iter_rows(min_row=CASE_VARS["STARTINGROW"], values_only=True, min_col=CASE_VARS["STARTINGCOL"]):
        client_id = row[CASE_VARS["CLIENTID"]]
        client = {
            'ID': clients_idx,
            'clientId': correct_legacy_id(client_id),
            'FirstName': row[CASE_VARS["FIRSTNAME"]],
            'LastName': row[CASE_VARS["LASTNAME"]],
            'Phone': clean_phone(row[CASE_VARS["PHONE"]]),
            'AddressNumber': row[CASE_VARS["ADDRESSNUMBER"]],
            'StreetName': row[CASE_VARS["STREETNAME"]],
            'ClientCity': row[CASE_VARS["CITY"]],
            'ClientState': row[CASE_VARS["STATE"]],
            'ClientCounty': row[CASE_VARS["COUNTY"]],
            'Zip': row[CASE_VARS["ZIP"]],
            'Race': correct_race(row[CASE_VARS["RACE"]]),
            'Ethnicity': correct_ethnicity(row[CASE_VARS["ETHNICITY"]]),
            'EnglishProficiencyLevel': correct_eng_prof(row[CASE_VARS["ENGLISHPROF"]]),
            'HouseholdIncome': row[CASE_VARS["HOUSEHOLDINCOME"]],
            'HouseholdSize': row[CASE_VARS["HOUSEHOLDSIZE"]],
            'ApartmentNumber': row[CASE_VARS["APTNUMBER"]],
            'Email': row[CASE_VARS["EMAIL"]],
            'InitialCaseType': correct_case_type(row[CASE_VARS["INITIALCASETYPE"]]),
            'DateOfBirth': correct_date(row[CASE_VARS["DATEOFBIRTH"]]),
            'CounselorName': row[CASE_VARS["COUNSELOR"]],
            'IntakeDate': correct_date(row[CASE_VARS["INTAKEDATE"]]),
            'CaseType': correct_case_type(row[CASE_VARS["CASETYPE"]]),
            'CaseStartDate': correct_date(row[CASE_VARS["CASESTARTDATE"]]),
            'CaseID': correct_case_id(row[CASE_VARS["CASEID"]]),
            'SessionID': row[CASE_VARS["SESSIONID"]],
            'Date': row[CASE_VARS["DATE"]],
            'NOFAGrant': correct_nofa(row[CASE_VARS["NOFAGRANT"]]),
            '9a': None,
            '9b': None,
            '9c': None,
            '9d': None,
            '9e': None,
            '9f': None,
            '10a': None,
            '10b': None,
            '10c': None,
            '10d': None,
            '10e': None,
            '10f': None,
            '10g': None,
            '10h': None,
            '10i': None,
            '10j': None,
            '10k': None,
            '10l': None,
            '10m': None,
            '10n': None,
            'SessionNotes': row[CASE_VARS["NOTES"]],
            'ruralareastatus': correct_rural(row[CASE_VARS["RURALSTATUS"]]),
            'HouseholdType': correct_household(row[CASE_VARS["HOUSEHOLDTYPE"]])
        }

        # Split the address
        if eval(FIX_ADDRESS.get()):
            split_address(row[CASE_VARS["STREETNAME"]], client)

        if eval(FIX_ZIPCODE.get()):
            get_zipcode(row[CASE_VARS["CITY"]], row[CASE_VARS["STATE"]], client)

        # Save the client by the ID for easy Access
        clients[clients_idx] = client
        clients_idx += 1

    # THIS IS DONE LAST FOR THE CLIENT SHEET
    # GRAB ALL DATA BEFORE THIS UNLESS TESTING
    # Add each client to the new spreadsheet
    for client in clients.values():
        template_client_sheet.append(list(client.values()))


def process_classes(class_data, vars):
    # Start creating initial hash table of classes
    classes = {}
    classes_idx = 1
    for row in class_data.active.iter_rows(min_row=vars["STARTINGROW"], values_only=True, min_col=vars["STARTINGCOL"]):
        client_id = row[vars["CLIENTID"]]
        # classDateFix = correct_date(row[17])
        client = {
            'ID': classes_idx,
            'clientId': correct_legacy_id(client_id),
            'FirstName': row[vars["FIRSTNAME"]],
            'LastName': row[vars["LASTNAME"]],
            'Phone': clean_phone(row[vars["PHONE"]]),
            'AddressNumber': row[vars["ADDRESSNUMBER"]],
            'StreetName': row[vars["STREETNAME"]],
            'ClientCity': row[vars["CITY"]],
            'ClientState': row[vars["STATE"]],
            'ClientCounty': row[vars["COUNTY"]],
            'Zip': row[vars["ZIP"]],
            'Race': correct_race(row[vars["RACE"]]),
            'Ethnicity': correct_ethnicity(row[vars["ETHNICITY"]]),
            'EnglishProficiencyLevel': correct_eng_prof(row[vars["ENGLISHPROF"]]),
            'HouseholdIncome': row[vars["HOUSEHOLDINCOME"]],
            'HouseholdSize': row[vars["HOUSEHOLDSIZE"]],
            'ApartmentNumber': row[vars["APTNUMBER"]],
            'Email': row[vars["EMAIL"]],
            'InitialCaseType': correct_case_type(row[vars["INITIALCASETYPE"]]),
            'DateOfBirth': correct_date(row[vars["DATEOFBIRTH"]]),
            'CounselorName': row[vars["COUNSELOR"]],
            'IntakeDate': correct_date(row[vars["INTAKEDATE"]]),
            'CourseID': row[vars["COURSEID"]],
            'ClassDate': row[vars["CLASSDATE"]],
            'AttendanceStatus': correct_attendance(row[vars["ATTENDANCESTATUS"]]),
            'RuralAreaStatus': correct_rural(row[vars["RURALSTATUS"]]),
            'HouseholdType': correct_household(row[vars["HOUSEHOLDTYPE"]])
        }

        if eval(FIX_ADDRESS.get()):
            split_address(row[CLASS_VARS["STREETNAME"]], client)

        if eval(FIX_ZIPCODE.get()):
            get_zipcode(row[CLASS_VARS["CITY"]], row[CLASS_VARS["STATE"]], client)

        # Save the client by the ID for easy Access
        classes[classes_idx] = client
        classes_idx += 1

    for client in classes.values():
        template_class_sheet.append(list(client.values()))


def browse_file():
    agency_filename = filedialog.askopenfilename()
    agency_data = load_workbook(filename=agency_filename, data_only=True)
    if eval(HAS_CLIENTS.get()):
        process_clients(agency_data)

    if eval(HAS_CLASSES.get()):
        if AGENCY.get() == 'MAHA':
            process_classes(agency_data, vars=MAHA_VARS)
        if AGENCY.get() == 'ABCDC':
            process_classes(agency_data, vars=ABCDC_VARS)

        process_classes(agency_data, vars=CLASS_VARS)


# Open Container
open_frame = customtkinter.CTkFrame(master=root, width=680, height=200, corner_radius=10)
open_frame.pack(padx=20, pady=20)

open_label = customtkinter.CTkLabel(open_frame, text="Select the Agency Spreadsheet to Open")
open_label.pack(padx=10, pady=10, side=tk.LEFT)

open_button = customtkinter.CTkButton(open_frame, text="Open Spreadsheet", command=browse_file)
open_button.pack(padx=10, pady=10, side=tk.RIGHT)

# Download Container
download_frame = customtkinter.CTkFrame(master=root, width=680, height=200, corner_radius=10)
download_frame.pack(padx=20, pady=20)
# create main entry and button

entry = customtkinter.CTkEntry(download_frame, placeholder_text="Save File As...", textvariable=SAVE_FILE_AS)
entry.pack(padx=20, pady=20)

download_button = customtkinter.CTkButton(download_frame, text="Download Spreadsheet", command=save_workbook)
download_button.pack(padx=10, pady=10)
downloaded_label = customtkinter.CTkLabel(download_frame, text="""Your file will be available in your downloads folder. Inspect your data
for any potential errors or data updates that need added to the script.
                                                                    """, font=('Arial', 14))
downloaded_label.pack(padx=10, pady=10)

# Activate mainloop for tkinter application
if __name__ == '__main__':
    root.mainloop()